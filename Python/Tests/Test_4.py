import psycopg2
from openpyxl import Workbook, load_workbook
import os
import decimal, datetime

connection_parameters = {
    "dbname": "test_3_db",
    "user": "postgres",
    "password": "newnew23@postgres",
    "host": "localhost",
    "port": "5432"
}

connection = psycopg2.connect(**connection_parameters)
cursor = connection.cursor()

wb = load_workbook("/home/shito/Learn_python/Tests/Test_3.xlsx") 

def convert_value(value):
    if isinstance(value, (int, decimal.Decimal)):
        return float(value)
    if isinstance(value, datetime.datetime):
        return value.date()  
    return value

for sheet in wb.sheetnames:
    ws = wb[f"{sheet}"]
    column_headers = [cell.value for cell in ws[1]]
    rows = list(ws.iter_rows(values_only=True))[1:]
    
    for row in rows:
        id = row[0]
        
        cursor.execute(f"SELECT * FROM {sheet} WHERE id = {id};")       # Get the specefic row by id from the database
        record = cursor.fetchone()
        
        for i in range(1, len(row)):                                    # Loop the row skipping the id 
            if convert_value(row[i]) != convert_value(record[i]):
                column_name = column_headers[i]
                if type(convert_value(row[i])) == str:
                    cursor.execute(f"""UPDATE {sheet}
                                    SET {column_name} = '{row[i]}'
                                    WHERE id = {id};""")
                else:
                    cursor.execute(f"""UPDATE {sheet}
                                    SET {column_name} = {row[i]}
                                    WHERE id = {id};""")
                connection.commit()     
                print(f"The record with the id = {id} from the table {sheet} was updated\nThe old value of {column_name} is : {record[i]}\nThe new value of {column_name} is : {row[i]}")
        
wb.save("Test_3.xlsx")

cursor.close()
connection.close()