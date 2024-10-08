import psycopg2
from openpyxl import Workbook, load_workbook
import os

connection_parameters = {
    "dbname": "test_3_db",
    "user": "postgres",
    "password": "newnew23@postgres",
    "host": "localhost",
    "port": "5432"
}

connection = psycopg2.connect(**connection_parameters)
cursor = connection.cursor()

cursor.execute("""
    SELECT table_name 
    FROM information_schema.tables 
    WHERE table_schema = 'public'
""")
tables = cursor.fetchall()

if os.path.exists('/home/shito/Learn_python/Tests/Test_3.xlsx'):
    wb = load_workbook("/home/shito/Learn_python/Tests/Test_3.xlsx") 
else:
    wb = Workbook()
    wb.remove(wb["Sheet"])
    
for table in tables:
    if table[0] not in wb.sheetnames:
        wb.create_sheet(f"{table[0]}")
    else:
        wb[f"{table[0]}"].delete_rows(1, wb[f"{table[0]}"].max_row)
    
    cursor.execute(f"SELECT * FROM {table[0]};")
    records = cursor.fetchall()
    columns_name = [desc[0] for desc in cursor.description]
    wb[f"{table[0]}"].append(columns_name)
    for record in records:
        wb[f"{table[0]}"].append(list(record))

for sheet in wb.sheetnames:
    delete = True
    for table in tables:
        if table[0] == sheet:
            delete = False
    if delete:
        wb.remove(wb[f"{sheet}"])

wb.save("Test_3.xlsx")

cursor.close()
connection.close()